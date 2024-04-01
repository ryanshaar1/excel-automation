import time
from datetime import datetime
from openpyxl import load_workbook

def main():
    start_time = None
    while True:
        input("Press Enter to start or stop the stopwatch: ")
        if start_time is None:
            start_time = datetime.now()
            print("Stopwatch started.")
        else:
            end_time = datetime.now()
            elapsed_time = end_time - start_time
            minutes = elapsed_time.total_seconds() // 60
            print(f"Stopwatch stopped. Total minutes elapsed: {minutes}")
            break

    wb = load_workbook('C:\\Users\\Ryan\\OneDrive\\מסמכים\\CODING\\Excel\\learning.xlsx')
    ws = wb.active

    ws['B2'].value = int(ws['B2'].value) + int(minutes)

    current_date = datetime.now().strftime("%Y-%m-%d")

    for row in range(1, ws.max_row + 1):
        if ws[f'A{row}'].value == current_date:
            ws[f'B{row}'].value = int(ws[f'B{row}'].value) + int(minutes)
            break
    else:
        ws.append([current_date, minutes])

    wb.save('C:\\Users\\Ryan\\OneDrive\\מסמכים\\CODING\\Excel\\learning.xlsx')

if __name__ == "__main__":
    main()