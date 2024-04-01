import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
from datetime import datetime




def main():
    start_time = None
    while True:
        input("Press Enter to start or stop the stopwatch: ")
        if start_time is None:
            start_time = datetime.now()
            print("Stopwatch started.")
        else:
            end_time = datetime.now()
            elapsed_time = end_time - start_time
            minutes = elapsed_time.total_seconds() // 60
            print(f"Stopwatch stopped. Total minutes elapsed: {minutes}")
            break
    

    #automation
    wb = load_workbook('C:\\Users\\Ryan\\OneDrive\\מסמכים\\CODING\\Excel\\learning.xlsx')
    ws = wb.active
    print(ws)
    ws['B2'].value = int(ws['B2'].value) + int(minutes)
   
    current_date = datetime.now().strftime("%Y-%m-%d")
    row = 37 #בתאריך 29.3.2024

    if ws['A{row}'].value == current_date:
        ws['B{row}'].value = int(ws['B{row}'].value) + int(minutes)
    else:
        row += 1
        ws['B{row}'].value = int(ws['B{row}'].value) + int(minutes)
    wb.save('C:\\Users\\Ryan\\OneDrive\\מסמכים\\CODING\\Excel\\learning.xlsx')
if __name__ == "__main__":
    main()
