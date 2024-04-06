import time
from datetime import datetime, timedelta
from openpyxl import load_workbook
import tkinter as tk

class StopwatchApp:
    def __init__(self, master):
        self.master = master
        self.start_time = None
        self.elapsed_time = None

        self.master.title("Stopwatch")
        self.master.configure(bg='black')

        screen_width = self.master.winfo_screenwidth()
        screen_height = self.master.winfo_screenheight()
        window_width = screen_width // 2
        window_height = screen_height // 2
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        self.master.geometry(f"{window_width}x{window_height}+{x}+{y}")

        self.label = tk.Label(master, text="Press the button to start or stop the stopwatch.", fg="white", bg="black", font=("Arial", 16))
        self.label.pack(pady=20)

        self.button = tk.Button(master, text="Start/Stop", command=self.toggle_stopwatch, font=("Arial", 20), bg="white", fg="black", width=10, height=2)
        self.button.pack()

        self.time_label = tk.Label(master, text="", fg="white", bg="black", font=("Arial", 16))
        self.time_label.pack(pady=20)

        self.message_entry = tk.Entry(master, font=("Arial", 14))
        self.message_entry.pack(pady=10)

        self.message_button = tk.Button(master, text="Add Message", command=self.add_message, font=("Arial", 16), bg="white", fg="black", width=12)
        self.message_button.pack(pady=10)

        self.update_time()

    def update_time(self):
        if self.start_time is not None:
            current_time = datetime.now()
            self.elapsed_time = current_time - self.start_time
            hours, remainder = divmod(self.elapsed_time.seconds, 3600)
            minutes, seconds = divmod(remainder, 60)
            self.time_label.config(text=f"Elapsed Time: {hours:02d}:{minutes:02d}:{seconds:02d}")
        self.master.after(1000, self.update_time)

    def toggle_stopwatch(self):
        if self.start_time is None:
            self.start_time = datetime.now()
            self.label.config(text="Stopwatch started.")
        else:
            end_time = datetime.now()
            self.elapsed_time = end_time - self.start_time
            minutes = self.elapsed_time.total_seconds() // 60
            self.label.config(text=f"Stopwatch stopped. Total minutes elapsed: {minutes}")
            self.update_excel(minutes)
            self.start_time = None

    def update_excel(self, minutes):
        wb = load_workbook('C:\\Users\\Ryan\\OneDrive\\מסמכים\\CODING\\Excel\\learning.xlsx')
        ws = wb.active

        ws['B2'].value = int(ws['B2'].value) + int(minutes)

        current_date = datetime.now().strftime("%Y-%m-%d")

        for row in range(1, ws.max_row + 1):
            if ws[f'A{row}'].value == current_date:
                ws[f'B{row}'].value = int(ws[f'B{row}'].value) + int(minutes)
                break
        else:
            ws.append([current_date, minutes])

        message = self.message_entry.get()
        if message:
            next_row = ws.max_row
            for row in range(1, ws.max_row + 1):
                if ws[f'C{row}'].value is None:
                    next_row = row
                    break
            ws[f'C{next_row}'].value = message

        wb.save('C:\\Users\\Ryan\\OneDrive\\מסמכים\\CODING\\Excel\\learning.xlsx')

    def add_message(self):
        message = self.message_entry.get()
        if message:
            self.message_entry.delete(0, tk.END)
            self.toggle_stopwatch()  # Stop the stopwatch if running
            if self.elapsed_time:  # Check if the stopwatch has been started
                elapsed_minutes = self.elapsed_time.total_seconds() // 60
                message += f" - {elapsed_minutes} minutes"
            self.update_excel_with_message(message)

    def update_excel_with_message(self, message):
        wb = load_workbook('C:\\Users\\Ryan\\OneDrive\\מסמכים\\CODING\\Excel\\learning.xlsx')
        ws = wb.active
        current_date = datetime.now().strftime("%Y-%m-%d")
        for row in range(1, ws.max_row + 1):
            if ws[f'A{row}'].value == current_date:
                max_column = ws.max_column
                for column in range(max_column, 0, -1):
                    if ws.cell(row=row, column=column).value is not None:
                        last_column_with_data = column
                        break
                else:
                    last_column_with_data = 1  # If there's no data in the row, start from column 1
                ws.cell(row=row, column=last_column_with_data + 1).value = message
                break
        wb.save('C:\\Users\\Ryan\\OneDrive\\מסמכים\\CODING\\Excel\\learning.xlsx')

def main():
    root = tk.Tk()
    app = StopwatchApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
